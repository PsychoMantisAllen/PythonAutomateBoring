import openpyxl
import os

os.chdir()  # change to the directory the excel file stored

workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)

sheet = workbook.get_sheet_by_name('sheet1')
type(sheet)

workbook.get_sheet_names()

cell = sheet['A1']     # getting the cell A1
cell.value

str(cell.value)        # changing to string

# alternatively we can do row and column to choose the cell
sheet.cell(row=1, column=2)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

